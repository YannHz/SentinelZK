"""
data/report_exporter.py
Constructor de los dos reportes Excel con formato profesional.
"""
import gc
import os
from datetime import date
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── Paleta de colores ──
COLOR_HEADER    = "1A1A2E"   # Azul oscuro
COLOR_PUNTUAL   = "D4EDDA"   # Verde claro
COLOR_TARDANZA  = "FFF3CD"   # Amarillo claro
COLOR_FALTA     = "F8D7DA"   # Rojo claro
COLOR_ALERTA    = "CCE5FF"   # Azul claro


def _aplicar_estilo_header(ws, fila: int, num_cols: int):
    """Aplica fondo oscuro y fuente blanca a la fila de encabezados."""
    fill   = PatternFill("solid", fgColor=COLOR_HEADER)
    fuente = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    alin   = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=fila, column=col)
        cell.fill   = fill
        cell.font   = fuente
        cell.alignment = alin


def _aplicar_borde(cell):
    lado = Side(style="thin", color="CCCCCC")
    cell.border = Border(left=lado, right=lado, top=lado, bottom=lado)


def _autofit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def exportar_reporte_a(df: pd.DataFrame, ruta_salida: str):
    """
    Reporte A: Logs crudos de auditoría.
    """
    if df.empty:
        print("[Exportador] Reporte A vacío, no se genera archivo.")
        return

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Logs Crudos", index=False, startrow=1)
        ws = writer.sheets["Logs Crudos"]

        # Título
        ws.cell(1, 1).value = "REPORTE A — LOGS DE AUDITORÍA"
        ws.cell(1, 1).font  = Font(bold=True, size=14, name="Calibri")

        _aplicar_estilo_header(ws, fila=2, num_cols=len(df.columns))
        _autofit_columns(ws)

        # Borde en datos
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=len(df.columns)):
            for cell in row:
                _aplicar_borde(cell)
                cell.alignment = Alignment(vertical="center")

    print(f"[Exportador] Reporte A generado: {ruta_salida}")


def exportar_reporte_b(df: pd.DataFrame, ruta_salida: str):
    """
    Reporte B: Consolidado directivo con colores por estado.
    """
    if df.empty:
        print("[Exportador] Reporte B vacío, no se genera archivo.")
        return

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Consolidado", index=False, startrow=1)
        ws = writer.sheets["Consolidado"]

        # Título
        ws.cell(1, 1).value = "REPORTE B — CONSOLIDADO DIRECTIVO"
        ws.cell(1, 1).font  = Font(bold=True, size=14, name="Calibri")

        _aplicar_estilo_header(ws, fila=2, num_cols=len(df.columns))
        _autofit_columns(ws)

        # Colorear filas según estado
        col_estado = list(df.columns).index("Estado") + 1 if "Estado" in df.columns else None

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=len(df.columns)):
            estado_cell = row[col_estado - 1] if col_estado else None
            estado_val  = estado_cell.value if estado_cell else ""

            color = None
            if estado_val == "Puntual":
                color = COLOR_PUNTUAL
            elif estado_val == "Tardanza":
                color = COLOR_TARDANZA
            elif estado_val == "Falta":
                color = COLOR_FALTA

            for cell in row:
                _aplicar_borde(cell)
                cell.alignment = Alignment(vertical="center")
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)

    gc.collect()
    print(f"[Exportador] Reporte B generado: {ruta_salida}")


def generar_reportes(df_a: pd.DataFrame, df_b: pd.DataFrame,
                     carpeta_salida: str = ".") -> None:
    """
    Genera ambos reportes y retorna las rutas de los archivos.
    """
    os.makedirs(carpeta_salida, exist_ok=True)
    hoy = date.today().strftime("%Y%m%d")

    ruta_a = os.path.join(carpeta_salida, f"SentinelZK_LogsCrudos_{hoy}.xlsx")
    ruta_b = os.path.join(carpeta_salida, f"SentinelZK_Consolidado_{hoy}.xlsx")

    exportar_reporte_a(df_a, ruta_a)
    exportar_reporte_b(df_b, ruta_b)

    return ruta_a, ruta_b
